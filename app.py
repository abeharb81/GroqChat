import streamlit as st
from groq import Groq
import tempfile, os, io, base64

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GroqChat",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap');

:root {
    --bg:       #0a0a0f;
    --surface:  #12121a;
    --border:   #2a2a3e;
    --accent:   #f97316;
    --accent2:  #fb923c;
    --text:     #e2e2f0;
    --muted:    #6b6b8a;
    --user-bg:  #1a1a2e;
    --bot-bg:   #111118;
    --glow:     0 0 24px rgba(249,115,22,0.25);
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}
#MainMenu, footer { visibility: hidden; }
.stDeployButton { display: none; }

[data-testid="collapsedControl"] {
    display: flex !important;
    visibility: visible !important;
    background: var(--accent) !important;
    border-radius: 0 8px 8px 0 !important;
    color: white !important;
    width: 28px !important;
    box-shadow: 4px 0 12px rgba(249,115,22,0.4) !important;
}

[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
    min-width: 280px !important;
}
[data-testid="stSidebar"] * { color: var(--text) !important; }

.sidebar-title {
    font-family: 'DM Serif Display', serif;
    font-size: 1.6rem;
    font-style: italic;
    color: var(--accent) !important;
    letter-spacing: -0.5px;
    margin-bottom: 2px;
}
.sidebar-sub {
    font-size: 0.7rem;
    color: var(--muted) !important;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    margin-bottom: 20px;
}

/* ── Nuclear textarea fix ── */
textarea, textarea:focus, textarea:active,
.stTextArea textarea,
.stTextArea > div > div > textarea,
div[data-baseweb="textarea"] textarea,
[data-testid="stSidebar"] textarea,
[class*="TextArea"] textarea {
    background-color: #1c1c30 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: 1px solid #3a3a55 !important;
    border-radius: 8px !important;
    font-size: 0.9rem !important;
    font-weight: 400 !important;
    line-height: 1.6 !important;
    opacity: 1 !important;
    caret-color: #f97316 !important;
}
textarea::placeholder {
    color: #5a5a7a !important;
    -webkit-text-fill-color: #5a5a7a !important;
    opacity: 1 !important;
}
textarea:focus {
    border-color: #f97316 !important;
    box-shadow: 0 0 0 2px rgba(249,115,22,0.2) !important;
    outline: none !important;
}
[data-testid="stSidebar"] input {
    background-color: #1c1c30 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border: 1px solid #3a3a55 !important;
    border-radius: 8px !important;
    opacity: 1 !important;
}

/* ── Main header ── */
.main-header {
    text-align: center;
    padding: 1.8rem 0 1rem;
}
.main-header h1 {
    font-family: 'DM Serif Display', serif;
    font-size: clamp(2rem, 5vw, 3rem);
    font-style: italic;
    background: linear-gradient(135deg, var(--accent), var(--accent2), #fbbf24);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    letter-spacing: -1px;
    margin: 0;
    line-height: 1.1;
}
.main-header p {
    font-size: 0.78rem;
    color: var(--muted);
    letter-spacing: 3px;
    text-transform: uppercase;
    margin-top: 6px;
}

/* ── Status badge ── */
.status-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: rgba(249,115,22,0.08);
    border: 1px solid rgba(249,115,22,0.25);
    border-radius: 999px;
    padding: 4px 14px;
    font-size: 0.7rem;
    color: var(--accent2);
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin: 0 auto 1.5rem;
    width: fit-content;
}
.status-dot {
    width: 6px; height: 6px;
    background: var(--accent);
    border-radius: 50%;
    animation: pulse 2s infinite;
}
@keyframes pulse {
    0%, 100% { opacity: 1; transform: scale(1); }
    50%       { opacity: 0.5; transform: scale(1.5); }
}

/* ── Input panels ── */
.input-panel {
    background: linear-gradient(135deg, #0f0f1a, #141420);
    border-radius: 14px;
    padding: 18px 20px;
    margin: 8px 0 14px;
}
.voice-panel  { border: 1px solid rgba(249,115,22,0.2); }
.file-panel   { border: 1px solid rgba(99,102,241,0.3); }

.panel-label {
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 10px;
}
.voice-panel .panel-label { color: #fb923c; }
.file-panel  .panel-label { color: #a5b4fc; }

.voice-transcript {
    background: #1a1a2e;
    border: 1px dashed #3a3a55;
    border-radius: 8px;
    padding: 10px 14px;
    font-size: 0.88rem;
    color: #c0c0e0;
    font-style: italic;
    margin-top: 10px;
    line-height: 1.5;
}

/* ── File badge shown in chat ── */
.file-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: rgba(99,102,241,0.12);
    border: 1px solid rgba(99,102,241,0.3);
    border-radius: 999px;
    padding: 3px 12px;
    font-size: 0.75rem;
    color: #a5b4fc;
    margin-bottom: 6px;
}

/* ── Active file indicator ── */
.active-file-bar {
    background: rgba(99,102,241,0.08);
    border: 1px solid rgba(99,102,241,0.25);
    border-radius: 10px;
    padding: 10px 16px;
    display: flex;
    align-items: center;
    gap: 10px;
    font-size: 0.82rem;
    color: #a5b4fc;
    margin-bottom: 12px;
}

/* ── Chat bubbles ── */
.user-bubble {
    background: var(--user-bg);
    border: 1px solid var(--border);
    border-radius: 16px 16px 4px 16px;
    padding: 14px 18px;
    margin: 8px 0 8px 40px;
    font-size: 0.93rem;
    line-height: 1.65;
}
.user-bubble.voice { border-left: 3px solid #f97316; }
.user-bubble.file  { border-left: 3px solid #6366f1; }
.bot-bubble {
    background: var(--bot-bg);
    border: 1px solid var(--border);
    border-left: 3px solid var(--accent);
    border-radius: 4px 16px 16px 16px;
    padding: 14px 18px;
    margin: 8px 40px 8px 0;
    font-size: 0.93rem;
    line-height: 1.65;
}

/* ── File uploader styling ── */
[data-testid="stFileUploader"] > div {
    background: #12121e !important;
    border: 1px dashed rgba(99,102,241,0.4) !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] p {
    color: #a5b4fc !important;
    font-size: 0.82rem !important;
}

/* ── Chat input ── */
[data-testid="stChatInput"] > div {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    box-shadow: var(--glow) !important;
}
[data-testid="stChatInput"] > div:focus-within { border-color: var(--accent) !important; }
[data-testid="stChatInput"] textarea {
    background: transparent !important;
    color: var(--text) !important;
    font-size: 0.93rem !important;
}

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, var(--accent), var(--accent2)) !important;
    border: none !important;
    border-radius: 8px !important;
    color: white !important;
    font-weight: 500 !important;
    transition: opacity 0.2s, transform 0.1s !important;
}
.stButton > button:hover { opacity: 0.85 !important; transform: translateY(-1px) !important; }

/* ── Stat cards ── */
.stat-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 12px;
    text-align: center;
}
.stat-num { font-family: 'DM Serif Display', serif; font-size: 1.6rem; color: var(--accent); line-height: 1; }
.stat-label { font-size: 0.65rem; color: var(--muted); letter-spacing: 2px; text-transform: uppercase; margin-top: 3px; }

.fancy-divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, var(--border), transparent);
    margin: 16px 0;
}

::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 2px; }
</style>
""", unsafe_allow_html=True)

# ── File extraction helpers ─────────────────────────────────────────────────
def extract_text_from_file(uploaded_file):
    """Extract text content from various file types."""
    name = uploaded_file.name
    ext  = name.rsplit(".", 1)[-1].lower()
    data = uploaded_file.read()

    # Plain text / code / markdown
    if ext in ("txt", "md", "py", "js", "ts", "html", "css", "json", "xml", "csv", "yaml", "yml", "sh", "sql"):
        try:
            return data.decode("utf-8")
        except UnicodeDecodeError:
            return data.decode("latin-1")

    # PDF
    elif ext == "pdf":
        try:
            import fitz  # PyMuPDF
            doc  = fitz.open(stream=data, filetype="pdf")
            text = "\n\n".join(page.get_text() for page in doc)
            doc.close()
            return text if text.strip() else "⚠️ This PDF appears to contain only images (scanned). Text extraction not possible."
        except ImportError:
            return "⚠️ PDF parsing library not available. Please install pymupdf."

    # Word .docx
    elif ext == "docx":
        try:
            from docx import Document
            doc  = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return "⚠️ DOCX parsing library not available. Please install python-docx."

    # Excel .xlsx / .xls
    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows.append(f"=== Sheet: {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(rows)
        except ImportError:
            return "⚠️ Excel parsing library not available. Please install openpyxl."

    # CSV (fallback in case not caught above)
    elif ext == "csv":
        try:
            return data.decode("utf-8")
        except Exception:
            return data.decode("latin-1")

    # Images — return base64 for vision model
    elif ext in ("png", "jpg", "jpeg", "gif", "webp"):
        b64 = base64.b64encode(data).decode()
        mime = f"image/{ext}" if ext != "jpg" else "image/jpeg"
        return f"__IMAGE__{mime}__SPLIT__{b64}"

    else:
        return f"⚠️ Unsupported file type: .{ext}"


def truncate(text, max_chars=12000):
    """Truncate file content so it fits in the context window."""
    if len(text) <= max_chars:
        return text
    half = max_chars // 2
    return text[:half] + f"\n\n... [truncated — {len(text)-max_chars:,} characters omitted] ...\n\n" + text[-half:]


# ── Session state ───────────────────────────────────────────────────────────
if "messages"       not in st.session_state: st.session_state.messages       = []
if "total_tokens"   not in st.session_state: st.session_state.total_tokens   = 0
if "message_count"  not in st.session_state: st.session_state.message_count  = 0
if "last_audio_id"  not in st.session_state: st.session_state.last_audio_id  = None
if "active_file"    not in st.session_state: st.session_state.active_file    = None  # {"name": ..., "content": ...}
if "last_file_name" not in st.session_state: st.session_state.last_file_name = None

# ── API Key from secrets ────────────────────────────────────────────────────
try:
    api_key = st.secrets["GROQ_API_KEY"]
except Exception:
    st.error("⚠️ API key not configured. Please add GROQ_API_KEY to your Streamlit secrets.")
    st.stop()

client = Groq(api_key=api_key)

# ── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sidebar-title">⚡ GroqChat</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-sub">Powered by Groq Cloud</div>', unsafe_allow_html=True)
    st.markdown('<div class="fancy-divider"></div>', unsafe_allow_html=True)

    model = st.selectbox(
        "🧠 Chat Model",
        options=[
            "llama-3.3-70b-versatile",
            "llama-3.1-8b-instant",
            "mixtral-8x7b-32768",
            "gemma2-9b-it",
        ],
        index=0,
        help="Llama 3.3 70B is recommended for file analysis.",
    )

    temperature = st.slider("🌡️ Temperature", 0.0, 2.0, 0.7, 0.05,
        help="Higher = more creative. Lower = more focused.")
    max_tokens = st.slider("📏 Max Tokens", 128, 8192, 2048, 128,
        help="Maximum length of the AI response.")

    system_prompt = st.text_area(
        "🎭 System Prompt",
        value="You are a helpful, knowledgeable, and friendly AI assistant. When given a file, analyze it carefully and answer questions about it accurately. Provide clear, structured, and thoughtful responses.",
        height=130,
        help="Shapes the AI's personality.",
    )

    st.markdown('<div class="fancy-divider"></div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""<div class="stat-card">
            <div class="stat-num">{st.session_state.message_count}</div>
            <div class="stat-label">Messages</div></div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""<div class="stat-card">
            <div class="stat-num">{st.session_state.total_tokens:,}</div>
            <div class="stat-label">Tokens</div></div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🗑️ Clear Conversation", use_container_width=True):
        st.session_state.messages       = []
        st.session_state.total_tokens   = 0
        st.session_state.message_count  = 0
        st.session_state.last_audio_id  = None
        st.session_state.active_file    = None
        st.session_state.last_file_name = None
        st.rerun()

# ── Main area ───────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>GroqChat</h1>
    <p>Voice · Files · Text · Cloud-powered AI</p>
</div>""", unsafe_allow_html=True)

st.markdown("""<div style="text-align:center">
    <div class="status-badge"><span class="status-dot"></span>Groq Cloud Connected</div>
</div>""", unsafe_allow_html=True)

st.markdown('<div class="fancy-divider"></div>', unsafe_allow_html=True)

# ── Chat history ────────────────────────────────────────────────────────────
for msg in st.session_state.messages:
    if msg["role"] == "user":
        icon        = "🎙️" if msg.get("voice") else ("📎" if msg.get("file") else "👤")
        extra_class = " voice" if msg.get("voice") else (" file" if msg.get("file") else "")
        content_html = ""
        if msg.get("file"):
            content_html += f'<div class="file-badge">📄 {msg["file"]}</div><br>'
        content_html += msg["content"]
        st.markdown(f'<div class="user-bubble{extra_class}">{icon} &nbsp;{content_html}</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="bot-bubble">⚡ &nbsp;{msg["content"]}</div>', unsafe_allow_html=True)

# ── Active file indicator ───────────────────────────────────────────────────
if st.session_state.active_file:
    col_f, col_x = st.columns([9, 1])
    with col_f:
        st.markdown(f"""<div class="active-file-bar">
            📄 <strong>{st.session_state.active_file['name']}</strong>
            &nbsp;—&nbsp; loaded &amp; ready · ask anything about it below
        </div>""", unsafe_allow_html=True)
    with col_x:
        if st.button("✕", help="Remove file", key="remove_file"):
            st.session_state.active_file    = None
            st.session_state.last_file_name = None
            st.rerun()

# ── Helper: send message ────────────────────────────────────────────────────
def send_message(prompt, is_voice=False, file_name=None, file_content=None):
    display_content = prompt
    st.session_state.messages.append({
        "role": "user", "content": prompt,
        "voice": is_voice, "file": file_name
    })

    icon        = "🎙️" if is_voice else ("📎" if file_name else "👤")
    extra_class = " voice" if is_voice else (" file" if file_name else "")
    html_content = ""
    if file_name:
        html_content += f'<div class="file-badge">📄 {file_name}</div><br>'
    html_content += display_content
    st.markdown(f'<div class="user-bubble{extra_class}">{icon} &nbsp;{html_content}</div>', unsafe_allow_html=True)

    # Build API messages — inject file content into the user message if provided
    api_messages = [{"role": "system", "content": system_prompt}]
    for m in st.session_state.messages[:-1]:
        api_messages.append({"role": m["role"], "content": m["content"]})

    # Final user message — attach file content if present
    if file_content and file_content.startswith("__IMAGE__"):
        # Vision: send as multipart content
        _, mime_b64 = file_content.split("__SPLIT__")
        mime = file_content.split("__IMAGE__")[1].split("__SPLIT__")[0]
        api_messages.append({
            "role": "user",
            "content": [
                {"type": "text",       "text": prompt},
                {"type": "image_url",  "image_url": {"url": f"data:{mime};base64,{mime_b64}"}}
            ]
        })
    elif file_content:
        combined = (
            f"I'm sharing a file called '{file_name}' with you.\n\n"
            f"--- FILE CONTENT START ---\n{truncate(file_content)}\n--- FILE CONTENT END ---\n\n"
            f"My question/request: {prompt}"
        )
        api_messages.append({"role": "user", "content": combined})
    else:
        # No new file — but if there's an active file in session, keep it in context
        active = st.session_state.active_file
        if active and not file_content:
            combined = (
                f"(Context: a file called '{active['name']}' has been loaded. "
                f"Refer to it if relevant.)\n\n"
                f"--- FILE CONTENT ---\n{truncate(active['content'])}\n---\n\n"
                f"{prompt}"
            )
            api_messages.append({"role": "user", "content": combined})
        else:
            api_messages.append({"role": "user", "content": prompt})

    try:
        with st.spinner("⚡ Thinking…"):
            response = client.chat.completions.create(
                model=model,
                messages=api_messages,
                temperature=temperature,
                max_tokens=max_tokens,
            )
        reply       = response.choices[0].message.content
        tokens_used = response.usage.total_tokens

        st.session_state.messages.append({"role": "assistant", "content": reply})
        st.session_state.total_tokens  += tokens_used
        st.session_state.message_count += 1

        st.markdown(f'<div class="bot-bubble">⚡ &nbsp;{reply}</div>', unsafe_allow_html=True)
        st.rerun()

    except Exception as e:
        st.error(f"❌ Error: {str(e)}")

# ══════════════════════════════════════════════════════════════════════════════
# INPUT PANELS — two columns side by side
# ══════════════════════════════════════════════════════════════════════════════
col_voice, col_file = st.columns(2)

# ── Voice panel ─────────────────────────────────────────────────────────────
with col_voice:
    st.markdown('<div class="input-panel voice-panel">', unsafe_allow_html=True)
    st.markdown('<div class="panel-label">🎙️ &nbsp;Voice Message</div>', unsafe_allow_html=True)

    audio_file = st.audio_input(
        label="Record",
        label_visibility="collapsed",
        key="voice_recorder",
    )

    if audio_file is not None:
        audio_id = id(audio_file)
        if audio_id != st.session_state.last_audio_id:
            st.session_state.last_audio_id = audio_id
            with st.spinner("🎙️ Transcribing…"):
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as tmp:
                        tmp.write(audio_file.read())
                        tmp_path = tmp.name
                    with open(tmp_path, "rb") as f:
                        transcription = client.audio.transcriptions.create(
                            model="whisper-large-v3-turbo",
                            file=("voice.wav", f, "audio/wav"),
                            response_format="text",
                        )
                    os.unlink(tmp_path)
                    transcript_text = (
                        transcription.strip()
                        if isinstance(transcription, str)
                        else transcription.text.strip()
                    )
                    if transcript_text:
                        st.markdown(f'<div class="voice-transcript">📝 {transcript_text}</div>', unsafe_allow_html=True)
                        send_message(transcript_text, is_voice=True)
                    else:
                        st.warning("Couldn't hear anything. Try again.")
                except Exception as e:
                    st.error(f"❌ Transcription error: {str(e)}")

    st.markdown('</div>', unsafe_allow_html=True)

# ── File upload panel ────────────────────────────────────────────────────────
with col_file:
    st.markdown('<div class="input-panel file-panel">', unsafe_allow_html=True)
    st.markdown('<div class="panel-label">📎 &nbsp;Upload a File</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        label="Upload",
        label_visibility="collapsed",
        type=["txt","md","py","js","ts","html","css","json","xml","csv",
              "yaml","yml","sh","sql","pdf","docx","xlsx","xls",
              "png","jpg","jpeg","gif","webp"],
        key="file_uploader",
        help="PDF, Word, Excel, CSV, images, code files, and more",
    )

    if uploaded is not None:
        if uploaded.name != st.session_state.last_file_name:
            st.session_state.last_file_name = uploaded.name
            with st.spinner(f"📖 Reading {uploaded.name}…"):
                content = extract_text_from_file(uploaded)
            st.session_state.active_file = {"name": uploaded.name, "content": content}
            st.success(f"✅ **{uploaded.name}** loaded! Now ask anything about it below.")

    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="fancy-divider"></div>', unsafe_allow_html=True)

# ── Text chat input ─────────────────────────────────────────────────────────
placeholder = (
    f"Ask about '{st.session_state.active_file['name']}' or anything else…"
    if st.session_state.active_file
    else "Type a message, or use voice / file upload above…"
)

if prompt := st.chat_input(placeholder):
    active = st.session_state.active_file
    if active:
        send_message(prompt, file_name=active["name"], file_content=active["content"])
        # After first message about file, keep file in context but don't re-attach full content every time
        st.session_state.active_file = {"name": active["name"], "content": active["content"]}
    else:
        send_message(prompt)
